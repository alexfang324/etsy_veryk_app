import re 
import openpyxl
import os
import base64
import PyPDF2

class DataProcessor:
###############################################################################
    @staticmethod
    #This function assume the first row of the file contains header information
    def readInFile(filename):
        fileExtension = filename.split('.')[1]
        dataHeader = []
        data=[] #data array to store copied and parsed data
        #############################################
        #handle Excel xlsx file with openpyxl library
        ############################################# 
        if fileExtension == 'xlsx':
            wb = openpyxl.load_workbook(filename)
            page = wb.active
            maxCol = page.max_column #get max column and row of the data in this worksheet
            maxRow = page.max_row
            
            #read in column header to array
            for j in range(1,maxCol+1):
                dataHeader.append(page.cell(1,j).value)

            #read in data to array
            for i in range(2,maxRow+1):
                row = []
                for j in range(1,maxCol+1):
                    row.append(page.cell(i,j).value)
                data.append(row)
        #############################################
        #handle csv file 
        #############################################     
        else:     
            with open(filename, 'r') as f:
                lines = f.read().split('\n') #create a list of rows of the file data
                firstLine = lines[0].replace('"','') #remove double quote sign from column header
                dataHeader = firstLine.split(',') #split column header into a list
                
                #save file data to a data array
                for line in lines[1:]:
                    if line == '': break #stop reading each row if we are at end of file
                    
                    #split line by comma only if it's not between quotes, this avoid parsing a product
                    #title into multiple parts
                    row = re.split(r",(?=(?:[^\"]*[\"][^\"]*[\"])*[^\"]*$)",line) 
                    data.append(row)
        return dataHeader, data
    
###############################################################################
    #This functions hands both Excel xlsx or csv write to file. It takes a
    #mandatory argument data which is an array of list(2D list) and optional
    #arguments inputFile, outputFile, and writeType. if no outputFile is specified,
    #data is written into inputFile by default. A valid call to this function is
    #like writeToFile(data,inputFile='name1.xlsx',outputFile='name2.xlsx',writeType='a')
    @staticmethod
    def writeToFile(data,**kwargs):
        #get input arguments
        inputFile = kwargs.get('inputFile',None)
        outputFile = kwargs.get('outputFile',None)
        writeType = kwargs.get('writeType',None)
        fileExtension = inputFile.split('.')[1]
        
        if not data:
            return
        #############################################
        #handle Excel xlsx file with openpyxl library
        #############################################
        if fileExtension == 'xlsx':
            if writeType == 'a':
                wb = openpyxl.load_workbook(inputFile)
            elif writeType == 'w':
                wb = openpyxl.workbook.Workbook()
            page = wb.active
            for row in data:
                page.append(row)
            if outputFile:
                wb.save(filename=outputFile) #save overwrites any existing data
            else:
                wb.save(filename=inputFile)          
                
        #############################################
        #handle csv file 
        #############################################     
        else:     
            #if output file name exist and write type is 'append' then copy old file data
            #to a new file first then write new data into it, otherwise we can just 
            # write to the appropriate file without copying transfering data
            if outputFile and writeType=='a':
                with open(inputFile,'r') as inf:
                    outf = open(outputFile,'w')
                    for line in inf:
                        outf.write(line)
                    f = outf
            elif outputFile:
                f = open(outputFile,writeType)
            else:
                f = open(inputFile,writeType)
                
            #write new data to file
            for i in range(len(data)):
                f.write(str(data[i][0])) #write data of first column
                for j in range(1,len(data[0])-1):
                    f.write(','+str(data[i][j])) #put a comma then write second column data
                f.write(','+str(data[i][j+1])+'\n') #switch to new line at the end of each row
            f.close()
            
###############################################################################
    #Using the information in Variations column, the function first find the total 
    #quantity sold, then combine rows with similar specs.
    #the Ind variables are the column index to the name, varaiations and quantity value
    #of the data array. unitWanted and specWanted provide keywords used to search
    #per pack quantity and product spec
    #Output Array = [Item Name, Specs, Total Quantity, Unit]
    @staticmethod
    def updateQuantityUsingVariations(data,nameInd,varInd,quanInd,unitWanted,specWanted):
        
        ###################################
        #find total quantity sold in each row and append value and units to the end of row
        ###################################
        for row in data:
            if not row[varInd]: #if Variation field is empty
                row.extend([row[quanInd], 'N/A']) #append original quantity with no unit
            else:
                for unit in unitWanted:
                    regex = '(\d)*[ ]*(?=dummy)' #regular expression to use for searching quantity
                    regex = re.sub('dummy',unit.lower(),regex) #replace dummy holder with proper unit
                    parsedNum = re.search(regex,row[varInd].lower()) #search expression in lower cases
                    if parsedNum: #if the per package quantitiy unit is found in Variation field
                        totQuantity = int(row[quanInd])*int(parsedNum[0])
                        row.extend([str(totQuantity),unit]) #append to end of row
                        break
                if not parsedNum: #if no per package quantity unit is found in Variations field
                    row.extend([row[quanInd],'N/A'])

             
        #############################
        #Combine rows with same specs in Variations column
        #############################
        accuData = {} #dict to accumulate data
        tQuanInd = len(data[0])-2
        if specWanted: #if specWanted list is not empty
            for row in data:
                key = [row[nameInd]]
                if not row[varInd]: #if Variation field is empty
                    key.extend(['' for spec in specWanted])
                else:
                    for spec in specWanted: #parse each spec
                        regex = '(?<=dummy:).*(?=[,-])' #regular expression to use for searching spec
                        regex = re.sub('dummy', spec.lower(),regex) #replace dummy holder with proper spec
                        parsedSpec = re.search(regex,row[varInd].lower()) #search expression in lower cases
                        if parsedSpec: #if the particular spec exist, add to key
                            key.append(spec+': '+parsedSpec[0]+' ')
                        else:
                            key.append('')
                key.append(data[data.index(row)][-1]) #append item units so we can use it when constructing output array
                key = tuple(key) #dict can't be list so we construct it with tuple        
                if key in accuData:
                    accuData[key] += int(row[tQuanInd])
                else:
                    accuData[key] = int(row[tQuanInd])

        #convert python dictionary back into array
        content = []
        for key, val in accuData.items():
            content.append([key[0],' '.join(key[1:-1]).strip(),str(val),key[-1]]) #strip() takes away white space on either end of string
        content.sort(key = lambda c: (c[0], c[1])) #sort by first column then sort by second column
        return content

###############################################################################
    #Uses the file product_name.csv to convert listing product name to a shorter
    #preferred name. The first column of the csv is the preferred name and the 
    #second column is the full listing name. The function input data is a 2D list
    #with the full listing name stored in the nameInd index of every row
    @staticmethod    
    def getPreferredName(data,nameInd,filename):
        
        #Generate Name Dict from file
        with open(filename,'r') as f:
            lines = f.read().split('\n') #create a list of rows of the file data
            nameDict = {} #dict to store name data
      
            #save file data to a data array
            for line in lines:
                if line == '': break #stop reading each row if we are at end of file
                
                #split line by comma only if it's not between quotes, this avoid parsing a product
                #title into multiple parts
                row = re.split(r",(?=(?:[^\"]*[\"][^\"]*[\"])*[^\"]*$)",line) 
                #if name not already in dict, save it to dict
                row = [item.strip() for item in row] #take away leading and trailing space of each string
                if row[1] not in nameDict:
                    nameDict[row[1]] = row[0]
                    
        #Convert names using name dict  
        for row in data:
            if row[nameInd] in nameDict:
                row[nameInd] = nameDict[row[nameInd]]
        return data

###############################################################################
    #Takes in a pdf output filename and an array of b64encoded content.
    #Existing output file will be overwritten to prevent lingering of historical data
    @staticmethod
    def createPDF(filename, contents):
        if not contents:
            print('No base64 encoded content given in DataProcessor.creatPDF()')
            return
        #if output file already exist, delete it first
        if os.path.isfile(filename):
           os.remove(filename) 
        
        # Decode the Base64 string, making sure that it contains only valid characters
        content = base64.b64decode(contents[0], validate=True)
        #check first 4 byte to see if it's for a pdf file       
        if content[0:4] != b'%PDF':
            raise ValueError('Missing the PDF file signature')
            
        #first loop, write data to a dummy output file object
        with open('temp.pdf', 'wb') as f:
            f.write(content)
        output = PyPDF2.PdfFileMerger() #output pdf object
        output.append(PyPDF2.PdfFileReader('temp.pdf', 'rb'))
        os.remove('temp.pdf') #delete dummy file
        
        #rest loop, append to output file object using similar approach
        for content in contents[1:]:
            content = base64.b64decode(content, validate=True)
            if content[0:4] != b'%PDF':
                raise ValueError('Missing the PDF file signature')
           
            #write new data to another dummy file
            with open('temp2.pdf', 'wb') as f:
                f.write(content)           
            output.append(PyPDF2.PdfFileReader('temp2.pdf', 'rb'))          
            os.remove('temp2.pdf')
        #write pdf object to file
        output.write(filename)
        output.close()

###############################################################################
    #Append input pdf to output pdf. If output file doesn't exist, input file will be 
    #renamed as the output file
    @staticmethod
    def appendPDF(inputfile,outputfile):
        if not os.path.isfile(inputfile):
            print(inputfile+' doens\'t exist')
            return
        if not inputfile.split('.')[1] == 'pdf':
            print('input file is not of pdf format. File is not appended')
            return
        
        #if outputfile doesn't exist, rename inputfile as outputfile
        if not os.path.isfile(outputfile):
            os.rename(inputfile,outputfile)
            return
        #otherwise, append to outputfile
        else:
            mergeFile = PyPDF2.PdfFileMerger()
            mergeFile.append(PyPDF2.PdfFileReader(outputfile, 'rb'))
            mergeFile.append(PyPDF2.PdfFileReader(inputfile, 'rb'))
            mergeFile.write(outputfile)

###############################################################################
    @staticmethod
    def deletePDF(filename):
        if os.path.isfile(filename):
            os.remove(filename)